"""Exportadores del snapshot a Excel y PDF (capa de presentación).

Funciones puras ``(ranking, meta, narrative) → bytes``: no leen disco ni red
(los datos llegan ya cargados del snapshot) y no dependen de Streamlit, así
que se testean solas. La app las conecta a botones de descarga pasándoles la
narrativa del idioma activo y el idioma (``lang``) para etiquetas y números.
"""

from io import BytesIO
from typing import Any, Final

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from tradefit import config
from tradefit.app.format import Lang, format_number, format_pct

#: Etiquetas de columna para los exports (subconjunto legible del ranking).
EXPORT_COLUMNS: Final[dict[Lang, dict[str, str]]] = {
    "es": {
        config.COL_RANK: "#",
        config.COL_COUNTRY: "ISO3",
        config.COL_COUNTRY_NAME: "Mercado",
        config.COL_MARKET_SIZE: "Importaciones prom. (USD)",
        config.COL_GROWTH: "Crecimiento (CAGR)",
        config.COL_SHARE: "Cuota del origen",
        config.COL_SHARE_TREND: "Δ cuota",
        config.COL_ORIGIN_EXPORT_SHARE: "% export. de Colombia",
        config.COL_COMPLEMENTARITY: "Complementariedad",
        config.COL_TARIFF: "Arancel enfrentado",
        config.COL_COMPETITOR_TARIFF: "Arancel competidores",
        config.COL_PREF_MARGIN: "Margen de preferencia",
        config.COL_LPI: "LPI logístico (1–5)",
        config.COL_DISTANCE_KM: "Distancia (km)",
        config.COL_ACCESSIBILITY: "Accesibilidad logística",
        config.COL_COVERAGE: "Cobertura de datos",
        config.COL_STABILITY: "Estabilidad macro",
        config.COL_SCORE: "Score bruto",
        config.COL_FINAL_SCORE: "Score final",
    },
    "en": {
        config.COL_RANK: "#",
        config.COL_COUNTRY: "ISO3",
        config.COL_COUNTRY_NAME: "Market",
        config.COL_MARKET_SIZE: "Avg. imports (USD)",
        config.COL_GROWTH: "Growth (CAGR)",
        config.COL_SHARE: "Origin's share",
        config.COL_SHARE_TREND: "Δ share",
        config.COL_ORIGIN_EXPORT_SHARE: "% of Colombia's exports",
        config.COL_COMPLEMENTARITY: "Complementarity",
        config.COL_TARIFF: "Tariff faced",
        config.COL_COMPETITOR_TARIFF: "Competitors' tariff",
        config.COL_PREF_MARGIN: "Preference margin",
        config.COL_LPI: "Logistics LPI (1–5)",
        config.COL_DISTANCE_KM: "Distance (km)",
        config.COL_ACCESSIBILITY: "Logistics accessibility",
        config.COL_COVERAGE: "Data coverage",
        config.COL_STABILITY: "Macro stability",
        config.COL_SCORE: "Raw score",
        config.COL_FINAL_SCORE: "Final score",
    },
}

#: Textos fijos de los exports, por idioma (mismo patrón que ``app/i18n.py``,
#: pero local: este módulo no depende de Streamlit ni de la sesión).
_STRINGS: Final[dict[str, dict[Lang, str]]] = {
    "sheet_narrative": {"es": "Narrativa", "en": "Narrative"},
    "recommendations": {
        "es": "Recomendación: dónde enfocarse",
        "en": "Recommendation: where to focus",
    },
    "final_score": {"es": "score final", "en": "final score"},
    "market_notes": {"es": "Lectura por mercado", "en": "Market notes"},
    "pdf_title": {
        "es": "Radar de Mercados — ranking de destinos",
        "en": "Market Radar — destination ranking",
    },
    "doc_title": {"es": "Radar de Mercados", "en": "Market Radar"},
    "ranking": {"es": "Ranking", "en": "Ranking"},
    "no_data": {"es": "s/d", "en": "n/a"},
    "meta_line": {
        "es": (
            "Producto: {label} · Origen: {origin} · Fuente: {source} · "
            "Datos {year_min}–{year_max} · RCA del origen: {rca}"
        ),
        "en": (
            "Product: {label} · Origin: {origin} · Source: {source} · "
            "Data {year_min}–{year_max} · Origin's RCA: {rca}"
        ),
    },
}

#: Cabecera de la tabla del PDF (subconjunto compacto), por idioma.
_PDF_HEADER: Final[dict[Lang, list[str]]] = {
    "es": ["#", "Mercado", "Import. prom. (USD M)", "CAGR", "Cuota", "Estab.", "Score final"],
    "en": ["#", "Market", "Avg. imports (USD M)", "CAGR", "Share", "Stab.", "Final score"],
}

#: Formato numérico Excel por columna del ranking.
_EXCEL_FORMATS: dict[str, str] = {
    config.COL_MARKET_SIZE: "#,##0",
    config.COL_GROWTH: "0.0%",
    config.COL_SHARE: "0.0%",
    config.COL_SHARE_TREND: "0.0%",
    config.COL_ORIGIN_EXPORT_SHARE: "0.0%",
    config.COL_COMPLEMENTARITY: "0.00",
    config.COL_TARIFF: "0.0%",
    config.COL_COMPETITOR_TARIFF: "0.0%",
    config.COL_PREF_MARGIN: "+0.0%;-0.0%",
    config.COL_LPI: "0.0",
    config.COL_DISTANCE_KM: "#,##0",
    config.COL_ACCESSIBILITY: "0.00",
    config.COL_COVERAGE: "0%",
    config.COL_STABILITY: "0.00",
    config.COL_SCORE: "0.000",
    config.COL_FINAL_SCORE: "0.000",
}

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(bold=True, color="FFFFFF")


def _meta_line(meta: dict[str, Any], lang: Lang) -> str:
    """Línea de contexto del snapshot (producto, origen, fuente, años)."""
    return _STRINGS["meta_line"][lang].format(
        label=meta.get("hs_label", ""),
        origin=meta.get("origin_iso3", ""),
        source=meta.get("source", ""),
        year_min=meta.get("data_year_min", ""),
        year_max=meta.get("data_year_max", ""),
        rca=meta.get("rca_balassa", "—"),
    )


def ranking_to_excel(
    ranking: pd.DataFrame,
    meta: dict[str, Any],
    narrative: dict[str, Any],
    lang: Lang = "es",
) -> bytes:
    """Arma el Excel del snapshot: hojas Ranking y Narrativa.

    Args:
        ranking: DataFrame conforme a ``ranking_schema``.
        meta: metadatos del snapshot (``meta.json``).
        narrative: narrativa de un idioma (``narrative.json[lang]``; puede
            ser vacía).
        lang: idioma de las etiquetas del archivo.

    Returns:
        Contenido del archivo ``.xlsx`` en bytes.
    """
    workbook = Workbook()

    # Snapshots construidos antes de una métrica nueva pueden no traer su
    # columna: se exporta solo lo presente.
    columns = {col: label for col, label in EXPORT_COLUMNS[lang].items() if col in ranking.columns}
    sheet = workbook.active
    sheet.title = _STRINGS["ranking"][lang]
    sheet.append(list(columns.values()))
    for cell in sheet[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for _, row in ranking.iterrows():
        sheet.append([row[col] if pd.notna(row[col]) else None for col in columns])
    for idx, col in enumerate(columns, start=1):
        letter = get_column_letter(idx)
        sheet.column_dimensions[letter].width = max(12, len(columns[col]) + 2)
        number_format = _EXCEL_FORMATS.get(col)
        if number_format:
            for cell in sheet[letter][1:]:
                cell.number_format = number_format
    sheet.freeze_panes = "A2"

    notes = workbook.create_sheet(_STRINGS["sheet_narrative"][lang])
    notes.column_dimensions["A"].width = 110
    notes.append([_meta_line(meta, lang)])
    notes.append([])
    recommendations = narrative.get("recommendations") or []
    if recommendations:
        notes.append([_STRINGS["recommendations"][lang]])
        notes["A3"].font = Font(bold=True, size=12)
        for i, rec in enumerate(recommendations, start=1):
            reasons = "; ".join(rec.get("reasons", []))
            score = format_number(float(rec.get("final_score", 0.0)), 3, lang)
            notes.append(
                [f"{i}. {rec.get('name')} ({_STRINGS['final_score'][lang]} {score}): {reasons}"]
            )
        notes.append([])
    markets = narrative.get("markets") or {}
    names = ranking.set_index(config.COL_COUNTRY)[config.COL_COUNTRY_NAME]
    for iso3 in ranking[config.COL_COUNTRY]:
        sentences = markets.get(iso3)
        if not sentences:
            continue
        header_row = notes.max_row + 1
        notes.append([f"{names.get(iso3, iso3)} ({iso3})"])
        notes[f"A{header_row}"].font = Font(bold=True)
        for sentence in sentences:
            notes.append([f"• {sentence}"])
        notes.append([])
    for row_cells in notes.iter_rows():
        for cell in row_cells:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def ranking_to_pdf(
    ranking: pd.DataFrame,
    meta: dict[str, Any],
    narrative: dict[str, Any],
    lang: Lang = "es",
) -> bytes:
    """Arma el PDF del snapshot: título, top-3, tabla y lectura por mercado.

    Args:
        ranking: DataFrame conforme a ``ranking_schema``.
        meta: metadatos del snapshot (``meta.json``).
        narrative: narrativa de un idioma (``narrative.json[lang]``; puede
            ser vacía).
        lang: idioma de las etiquetas del archivo.

    Returns:
        Contenido del archivo ``.pdf`` en bytes.
    """
    styles = getSampleStyleSheet()
    body = ParagraphStyle("body", parent=styles["Normal"], fontSize=9, leading=12)
    h2 = styles["Heading2"]
    h3 = styles["Heading3"]

    story: list[Any] = [
        Paragraph(_STRINGS["pdf_title"][lang], styles["Title"]),
        Paragraph(_meta_line(meta, lang), body),
        Spacer(1, 0.4 * cm),
    ]

    recommendations = narrative.get("recommendations") or []
    if recommendations:
        story.append(Paragraph(_STRINGS["recommendations"][lang], h2))
        for i, rec in enumerate(recommendations, start=1):
            reasons = "; ".join(rec.get("reasons", []))
            score = format_number(float(rec.get("final_score", 0.0)), 3, lang)
            story.append(
                Paragraph(
                    f"<b>{i}. {rec.get('name')}</b> "
                    f"({_STRINGS['final_score'][lang]} {score}): {reasons}",
                    body,
                )
            )
        story.append(Spacer(1, 0.4 * cm))

    story.append(Paragraph(_STRINGS["ranking"][lang], h2))
    rows: list[list[str]] = [_PDF_HEADER[lang]]
    no_data = _STRINGS["no_data"][lang]
    for _, row in ranking.iterrows():
        growth = row[config.COL_GROWTH]
        rows.append(
            [
                str(int(row[config.COL_RANK])),
                str(row[config.COL_COUNTRY_NAME]),
                format_number(row[config.COL_MARKET_SIZE] / 1e6, 0, lang),
                no_data if pd.isna(growth) else format_pct(float(growth), 1, lang),
                format_pct(float(row[config.COL_SHARE]), 1, lang),
                format_number(float(row[config.COL_STABILITY]), 2, lang),
                format_number(float(row[config.COL_FINAL_SCORE]), 3, lang),
            ]
        )
    table = Table(rows, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ALIGN", (2, 1), (-1, -1), "RIGHT"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F6FA")]),
            ]
        )
    )
    story.append(table)

    markets = narrative.get("markets") or {}
    if markets:
        story.append(Spacer(1, 0.5 * cm))
        story.append(Paragraph(_STRINGS["market_notes"][lang], h2))
        names = ranking.set_index(config.COL_COUNTRY)[config.COL_COUNTRY_NAME]
        for iso3 in ranking[config.COL_COUNTRY]:
            sentences = markets.get(iso3)
            if not sentences:
                continue
            story.append(Paragraph(f"{names.get(iso3, iso3)} ({iso3})", h3))
            for sentence in sentences:
                story.append(Paragraph(f"• {sentence}", body))

    buffer = BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        title=_STRINGS["doc_title"][lang],
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )
    document.build(story)
    return buffer.getvalue()
