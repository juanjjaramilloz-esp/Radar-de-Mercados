"""Tests de los exportadores Excel/PDF (sin Streamlit, sin disco)."""

from io import BytesIO
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from tradefit import config
from tradefit.app.export import ranking_to_excel, ranking_to_pdf


def _ranking() -> pd.DataFrame:
    return pd.DataFrame(
        {
            config.COL_RANK: [1, 2],
            config.COL_COUNTRY: ["AAA", "BBB"],
            config.COL_COUNTRY_NAME: ["Alfalandia", "Betalandia"],
            config.COL_MARKET_SIZE: [2_000_000_000.0, 500_000_000.0],
            config.COL_GROWTH: [0.10, float("nan")],
            config.COL_SHARE: [0.25, 0.0],
            config.COL_SHARE_TREND: [0.02, 0.0],
            config.COL_COMPLEMENTARITY: [0.8, 0.1],
            config.COL_RCA: [9.0, 9.0],
            config.COL_STABILITY: [0.9, 0.4],
            config.COL_SCORE: [0.8, 0.3],
            config.COL_FINAL_SCORE: [0.76, 0.21],
        }
    )


def _meta() -> dict[str, Any]:
    return {
        "hs_label": "Café (HS 0901)",
        "origin_iso3": "COL",
        "source": "stub",
        "data_year_min": 2022,
        "data_year_max": 2024,
        "rca_balassa": 9.0,
    }


def _narrative() -> dict[str, Any]:
    return {
        "recommendations": [
            {
                "iso3": "AAA",
                "name": "Alfalandia",
                "final_score": 0.76,
                "reasons": ["1.º destino por tamaño de mercado (USD 2.000 M/año)"],
            }
        ],
        "markets": {
            "AAA": ["Importa USD 2.000 M al año del producto."],
            "BBB": ["Importa USD 500 M al año del producto."],
        },
    }


def test_excel_se_reabre_con_las_celdas_esperadas() -> None:
    data = ranking_to_excel(_ranking(), _meta(), _narrative())
    workbook = load_workbook(BytesIO(data))
    assert workbook.sheetnames == ["Ranking", "Narrativa"]
    sheet = workbook["Ranking"]
    assert sheet["C1"].value == "Mercado"
    assert sheet["C2"].value == "Alfalandia"
    assert sheet["D2"].value == 2_000_000_000.0
    # El NaN de crecimiento de BBB queda como celda vacía, no como texto "nan"
    assert sheet["E3"].value is None
    narrativa = workbook["Narrativa"]
    texts = [str(c.value) for row in narrativa.iter_rows() for c in row if c.value]
    assert any("Alfalandia" in t for t in texts)


def test_excel_sin_narrativa_no_rompe() -> None:
    data = ranking_to_excel(_ranking(), _meta(), {})
    workbook = load_workbook(BytesIO(data))
    assert "Ranking" in workbook.sheetnames


def test_excel_en_ingles_traduce_etiquetas() -> None:
    data = ranking_to_excel(_ranking(), _meta(), _narrative(), lang="en")
    workbook = load_workbook(BytesIO(data))
    assert workbook.sheetnames == ["Ranking", "Narrative"]
    sheet = workbook["Ranking"]
    assert sheet["C1"].value == "Market"
    narrative = workbook["Narrative"]
    texts = [str(c.value) for row in narrative.iter_rows() for c in row if c.value]
    assert any("Recommendation: where to focus" in t for t in texts)
    assert any("Product: Café (HS 0901)" in t for t in texts)  # meta line en inglés


def test_pdf_en_ingles_es_valido() -> None:
    data = ranking_to_pdf(_ranking(), _meta(), _narrative(), lang="en")
    assert data.startswith(b"%PDF")


def test_pdf_es_un_pdf_valido() -> None:
    data = ranking_to_pdf(_ranking(), _meta(), _narrative())
    assert data.startswith(b"%PDF")
    assert len(data) > 1000  # documento no trivial (título + tabla + narrativa)


def test_pdf_sin_narrativa_no_rompe() -> None:
    data = ranking_to_pdf(_ranking(), _meta(), {})
    assert data.startswith(b"%PDF")
